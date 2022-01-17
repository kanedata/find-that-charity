import datetime
import io

from openpyxl import load_workbook

from ftc.management.commands._base_scraper import HTMLScraper
from ftc.models import Organisation, OrganisationLocation


class Command(HTMLScraper):
    """
    Spider for scraping details of Registered Social Landlords in England
    """

    name = "rsl"
    allowed_domains = ["gov.uk", "githubusercontent.com"]
    start_urls = [
        "https://www.gov.uk/government/publications/current-registered-providers-of-social-housing",
    ]
    org_id_prefix = "GB-SHPE"
    id_field = "registration number"
    source = {
        "title": "Current registered providers of social housing",
        "description": "Current registered providers of social housing and new registrations and deregistrations. Covers England",
        "identifier": "rsl",
        "license": "http://www.nationalarchives.gov.uk/doc/open-government-licence/version/3/",
        "license_name": "Open Government Licence v3.0",
        "issued": "",
        "modified": "",
        "publisher": {
            "name": "Regulator of Social Housing",
            "website": "https://www.gov.uk/government/organisations/regulator-of-social-housing",
        },
        "distribution": [
            {
                "downloadURL": "",
                "accessURL": "",
                "title": "Current registered providers of social housing",
            }
        ],
    }
    orgtypes = ["Registered Provider of Social Housing"]

    def parse_file(self, response, source_url):
        link = [link for link in response.html.links if link.endswith(".xlsx")][0]
        self.set_download_url(link)
        r = self.session.get(link)
        r.raise_for_status()

        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        sheets = [
            sheetname
            for sheetname in wb.sheetnames
            if "listing" in sheetname.lower() or "find view" in sheetname.lower()
        ]
        for sheetname in sheets:
            ws = wb[sheetname]

            # self.source["issued"] = wb.properties.modified.isoformat()[0:10]

            headers = None
            for k, row in enumerate(ws.rows):
                if not headers:
                    headers = [c.value.lower() for c in row]
                else:
                    record = dict(zip(headers, [c.value for c in row]))
                    self.parse_row(record)

    def parse_row(self, record):

        record = self.clean_fields(record)
        if not record.get("organisation name") or not record.get("registration number"):
            return

        org_types = [
            self.add_org_type("Registered Provider of Social Housing"),
        ]
        if record.get("corporate form"):
            if record["corporate form"] == "Company":
                org_types.append(self.add_org_type("Registered Company"))
                org_types.append(
                    self.add_org_type(
                        "{} {}".format(record["designation"], record["corporate form"])
                    )
                )
            elif record["corporate form"] == "CIO-Charitable Incorporated Organisation":
                org_types.append(
                    self.add_org_type("Charitable Incorporated Organisation")
                )
                org_types.append(self.add_org_type("Registered Charity"))
            elif record["corporate form"] == "Charitable Company":
                org_types.append(self.add_org_type("Registered Company"))
                org_types.append(self.add_org_type("Incorporated Charity"))
                org_types.append(self.add_org_type("Registered Charity"))
            elif record["corporate form"] == "Unincorporated Charity":
                org_types.append(self.add_org_type("Registered Charity"))
            else:
                org_types.append(self.add_org_type(record["corporate form"]))
        elif record.get("designation"):
            org_types.append(self.add_org_type(record["designation"]))

        org_ids = [self.get_org_id(record)]

        self.add_org_record(
            Organisation(
                **{
                    "org_id": self.get_org_id(record),
                    "name": record.get("organisation name"),
                    "charityNumber": None,
                    "companyNumber": None,
                    "streetAddress": None,
                    "addressLocality": None,
                    "addressRegion": None,
                    "addressCountry": "England",
                    "postalCode": None,
                    "telephone": None,
                    "alternateName": [],
                    "email": None,
                    "description": None,
                    "organisationType": [o.slug for o in org_types],
                    "organisationTypePrimary": org_types[0],
                    "url": None,
                    # "location": locations,
                    "latestIncome": None,
                    "dateModified": datetime.datetime.now(),
                    "dateRegistered": record.get("registration date"),
                    "dateRemoved": None,
                    "active": True,
                    "parent": None,
                    "orgIDs": org_ids,
                    "scrape": self.scrape,
                    "source": self.source,
                    "spider": self.name,
                    "org_id_scheme": self.orgid_scheme,
                }
            )
        )
