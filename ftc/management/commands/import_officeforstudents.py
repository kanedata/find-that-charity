import datetime
import io

from django.utils.text import slugify
from openpyxl import load_workbook

from ftc.management.commands._base_scraper import BaseScraper
from ftc.models import Organisation


class Command(BaseScraper):
    name = "officeforstudents"
    allowed_domains = ["officeforstudents.org.uk"]
    start_urls = ["https://register-api.officeforstudents.org.uk/api/Download/"]
    org_id_prefix = "GB-UKPRN"
    id_field = "providers-ukprn"
    source = {
        "title": "The OfS Register",
        "description": """The Register lists all the English higher education providers registered by the OfS.

It is a single, authoritative reference about a provider’s regulatory status.""",
        "identifier": "officeforstudents",
        "license": "http://www.nationalarchives.gov.uk/doc/open-government-licence/version/3/",
        "license_name": "Open Government Licence v3.0",
        "issued": "",
        "modified": "",
        "publisher": {
            "name": "Office for Students",
            "website": "https://www.officeforstudents.org.uk/",
        },
        "distribution": [
            {
                "downloadURL": "https://register-api.officeforstudents.org.uk/api/Download/",
                "accessURL": "https://www.officeforstudents.org.uk/for-providers/regulatory-resources/the-ofs-register/",
                "title": "The OfS Register",
            }
        ],
    }
    orgtypes = [
        "Higher education institution",
        "University",
        "Registered Charity",
        "Exempt Charity",
        "Registered Charity (Scotland)",
    ]

    def parse_file(self, response, source_url):
        wb = load_workbook(io.BytesIO(response.content), read_only=True)
        # for sheet in wb.sheetnames:
        for sheet in ["Register"]:
            headers = []
            for k, row in enumerate(wb[sheet].rows):
                if not row[1].value:
                    continue

                if not headers:
                    headers = [
                        slugify(c.value.lower().strip()) if c.value else f"column-{i}"
                        for i, c in enumerate(row)
                    ]
                else:
                    record = dict(zip(headers, [c.value for c in row]))
                    record["sheet"] = sheet
                    self.parse_row(record)

    def parse_row(self, record):
        record = self.clean_fields(record, blank_values=["", "Not applicable"])

        altnames = []
        altname_fields = [
            "embedded-colleges-covered-by-the-providers-registration",
            "providers-trading-names",
        ]
        for f in altname_fields:
            if record.get(f):
                altnames.extend(record[f].split("\n"))

        address = record.get("providers-contact-address")
        new_address = []

        streetAddress = None
        addressLocality = None
        addressRegion = None
        addressCountry = None
        postalCode = None
        if address:
            address = address.split("\n")
            for address_line in address:
                address_line = address_line.strip()
                if self.postcode_regex.match(address_line):
                    postalCode = address_line
                elif address_line in ("United Kingdom", "United Kindgom"):
                    addressCountry = "United Kingdom"
                else:
                    new_address.append(address_line)

            if len(new_address) > 1:
                addressLocality = new_address[-1]
                streetAddress = ", ".join(new_address[:-1])
            else:
                streetAddress = new_address[0]

        orgtypes = [
            self.orgtype_cache["higher-education-institution"],
        ]
        if record.get("is-the-provider-an-exempt-or-registered-charity") in (
            "Registered",
            "Regstered",
            "Registered as a charity in Scotland",
        ):
            orgtypes.append(self.orgtype_cache["registered-charity"])
        if record.get("is-the-provider-an-exempt-or-registered-charity") in (
            "Registered as a charity in Scotland",
        ):
            orgtypes.append(self.orgtype_cache["registered-charity-scotland"])
        if (
            record.get(
                "does-the-provider-have-the-right-to-use-university-in-its-title"
            )
            == "Yes"
        ):
            orgtypes.append(self.orgtype_cache["university"])

        result = {
            "org_id": self.get_org_id(record),
            "name": record.get("providers-legal-name"),
            "charityNumber": None,
            "companyNumber": None,
            "streetAddress": streetAddress,
            "addressLocality": addressLocality,
            "addressRegion": addressRegion,
            "addressCountry": addressCountry,
            "postalCode": postalCode,
            "telephone": None,
            "alternateName": altnames,
            "email": record.get("providers-email-address"),
            "description": None,
            "organisationType": [o.slug for o in orgtypes],
            "organisationTypePrimary": orgtypes[0],
            "url": record.get("providers-website"),
            "latestIncome": None,
            "dateModified": datetime.datetime.now(),
            "dateRegistered": None,
            "dateRemoved": None,
            "active": True,
            "parent": None,
            "orgIDs": [self.get_org_id(record)],
            "scrape": self.scrape,
            "source": self.source,
            "spider": self.name,
            "org_id_scheme": self.orgid_scheme,
        }
        self.add_org_record(Organisation(**result))
