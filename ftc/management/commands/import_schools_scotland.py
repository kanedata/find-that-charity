# -*- coding: utf-8 -*-
import datetime
import io

from openpyxl import load_workbook

from ftc.management.commands._base_scraper import AREA_TYPES, HTMLScraper
from ftc.models import Organisation

SCOT_LAS = {
    "Aberdeen City": "S12000033",
    "Aberdeenshire": "S12000034",
    "Angus": "S12000041",
    "Argyll & Bute": "S12000035",
    "Clackmannanshire": "S12000005",
    "Dumfries & Galloway": "S12000006",
    "Dundee City": "S12000042",
    "East Ayrshire": "S12000008",
    "East Dunbartonshire": "S12000045",
    "East Lothian": "S12000010",
    "East Renfrewshire": "S12000011",
    "Edinburgh City": "S12000036",
    "Falkirk": "S12000014",
    "Fife": "S12000015",
    "Glasgow City": "S12000046",
    "Highland": "S12000017",
    "Inverclyde": "S12000018",
    "Midlothian": "S12000019",
    "Moray": "S12000020",
    "Na h-Eileanan Siar": "S12000013",
    "North Ayrshire": "S12000021",
    "North Lanarkshire": "S12000044",
    "Orkney Islands": "S12000023",
    "Perth & Kinross": "S12000024",
    "Renfrewshire": "S12000038",
    "Scottish Borders": "S12000026",
    "Shetland Islands": "S12000027",
    "South Ayrshire": "S12000028",
    "South Lanarkshire": "S12000029",
    "Stirling": "S12000030",
    "West Dunbartonshire": "S12000039",
    "West Lothian": "S12000040",
}


class Command(HTMLScraper):
    name = "schools_scotland"
    allowed_domains = ["gov.scot"]
    start_urls = [
        "http://www.gov.scot/Topics/Statistics/Browse/School-Education/Datasets/contactdetails"
    ]
    skip_rows = 5
    org_id_prefix = "GB-SCOTEDU"
    id_field = "seedcode"
    source = {
        "title": "School Contact Details",
        "description": "School contact details as at September 2017 including school names, addresses, pupil rolls, FTE numbers of teachers, urban/rural classification, denomination and proportion of pupils from minority ethnic groups.",
        "identifier": "schoolsscotland",
        "license": "http://www.nationalarchives.gov.uk/doc/open-government-licence/",
        "license_name": "Open Government Licence",
        "issued": "",
        "modified": "",
        "publisher": {
            "name": "Scottish Government",
            "website": "https://www.gov.scot/",
        },
        "distribution": [
            {"downloadURL": "", "accessURL": "", "title": "School Contact Details"}
        ],
    }
    orgtypes = ["Education"]

    def parse_file(self, response, source_url):
        link = [
            link for link in response.html.absolute_links if link.endswith(".xlsx")
        ][0]
        self.set_download_url(link)
        r = self.session.get(link)

        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        latest_sheet = wb[
            sorted([s for s in wb.sheetnames if s.startswith("Open at")])[-1]
        ]

        # self.source["issued"] = wb.properties.modified.isoformat()[0:10]

        self.logger.info("Latest sheet: {}".format(latest_sheet.title))
        headers = {}
        seen_blank_row = False
        for k, row in enumerate(latest_sheet.rows):
            if k < self.skip_rows or seen_blank_row:
                continue
            elif k == self.skip_rows:
                headers = self.get_headers(row, latest_sheet, k)
                continue
            elif row[0].value is None:
                seen_blank_row = True
                continue

            record = {}
            for i, c in enumerate(row):
                if i + 1 in headers:
                    v = c.value
                    if v in ["", ".", "N/A", "0", 0]:
                        v = None
                    record[headers[i + 1]] = v

            org_id = self.get_org_id(record)
            org_types = self.get_org_types(record)

            self.records.append(
                Organisation(
                    org_id=org_id,
                    name=record.get("school_name"),
                    charityNumber=None,
                    companyNumber=None,
                    streetAddress=record.get("address_1"),
                    addressLocality=record.get("address_2"),
                    addressRegion=record.get("address_3"),
                    addressCountry="Scotland",
                    postalCode=self.parse_postcode(record.get("post_code")),
                    telephone=record.get("phone"),
                    alternateName=[],
                    email=record.get("e_mail"),
                    description=None,
                    organisationType=[o.slug for o in org_types],
                    organisationTypePrimary=org_types[0],
                    url=None,
                    location=self.get_locations(record),
                    latestIncome=None,
                    dateModified=datetime.datetime.now(),
                    dateRegistered=None,
                    dateRemoved=None,
                    active=True,
                    parent=None,
                    orgIDs=[org_id],
                    scrape=self.scrape,
                    source=self.source,
                    spider=self.name,
                    org_id_scheme=self.orgid_scheme,
                )
            )

    def get_headers(self, row, sheet, row_number):
        previous_overtitle = None
        header_names = []
        for c in row:
            if c.value:
                title = str(c.value)

                # get the row before to find the heading for this title
                overtitle = sheet.cell(row_number, c.column).value
                if overtitle is None:
                    overtitle = previous_overtitle
                else:
                    previous_overtitle = overtitle

                # we actually only need three of them
                if overtitle:
                    if overtitle.startswith("Pupil rolls"):
                        overtitle = "Pupil rolls"
                    elif overtitle.startswith("Teachers"):
                        overtitle = "Teachers FTE"
                    elif overtitle.startswith("School type"):
                        overtitle = "School type"
                    else:
                        overtitle = None

                header_names.append(
                    self.slugify("{} {}".format(overtitle if overtitle else "", title))
                )

        return dict(
            zip(
                [c.column for c in row if c.value],  # header column numbers
                header_names,
            )
        )

    def get_org_types(self, record):
        org_types = [
            self.orgtype_cache["education"],
            self.add_org_type(record.get("centre_type") + " School"),
        ]
        for f in [
            "school_type_primary",
            "school_type_secondary",
            "school_type_special",
        ]:
            if record.get(f):
                org_types.append(self.add_org_type(record[f] + " School"))
        return org_types

    def get_locations(self, record):
        locations = []
        if SCOT_LAS.get(record.get("la_name")):
            code = SCOT_LAS.get(record.get("la_name"))
            locations.append(
                {
                    "id": code,
                    "name": record.get("la_name"),
                    "geoCode": code,
                    "geoCodeType": AREA_TYPES.get(code[0:3], "Local Authority"),
                }
            )
        return locations
